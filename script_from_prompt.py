#!/usr/bin/env python3
import os
import asyncio
import aiohttp
from bs4 import BeautifulSoup
import xlsxwriter
from openpyxl import load_workbook
from tqdm.asyncio import tqdm
import logging
import re

# ── CONFIGURATION ────────────────────────────────────────────────────────────
BASE_URL            = "https://natiga.mansourals.com/Elgohary/sec26_t1/index.php"
GRADE               = "الثاني الثانوي"
ACTION              = "search"
START_SEAT          = 261370     # where you want to resume
END_SEAT            = 55602
CONCURRENT_REQUESTS = 200
RETRIES             = 2
TIMEOUT             = 10
OUTPUT_XLSX         = "grades.xlsx"
ERROR_LOG           = "errors.txt"
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    filename=ERROR_LOG,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s"
)

# --- 1) Load existing data if any ---
existing = []
seen_codes = set()
if os.path.exists(OUTPUT_XLSX):
    wb_old = load_workbook(OUTPUT_XLSX, read_only=True)
    ws_old = wb_old.active
    header_row = next(ws_old.iter_rows(min_row=1, max_row=1, values_only=True))
    # Find the index of the code column (case-insensitive)
    code_col = None
    for idx, h in enumerate(header_row):
        if h and str(h).strip().lower() == "code":
            code_col = idx
            break
    if code_col is None:
        print("⚠️  Could not find a 'code' column in existing file—resuming all seats.")
    else:
        for row in ws_old.iter_rows(min_row=2, values_only=True):
            # guard against short rows
            if code_col >= len(row):
                continue
            code_val = row[code_col]
            if code_val is None:
                continue
            try:
                seat_str = str(int(code_val)).zfill(6)
            except:
                continue
            seen_codes.add(seat_str)
            existing.append(dict(zip(header_row, row)))
        print(f"Loaded {len(existing)} existing records, skipping those seats.")
    wb_old.close()

# --- 2) Prepare async fetcher ---
semaphore = asyncio.Semaphore(CONCURRENT_REQUESTS)
results = []

async def fetch_one(session, seat_str):
    params = {"stdInfo": seat_str, "grade": GRADE, "action": ACTION}
    async with semaphore:
        for _ in range(RETRIES):
            try:
                async with session.get(BASE_URL, params=params, timeout=TIMEOUT) as resp:
                    text = await resp.text()
                soup = BeautifulSoup(text, "html.parser")
                sec = soup.find("section", id="student_result")
                if not sec:
                    return None
                # parse name
                name_tag = sec.find("h2")
                name = name_tag.get_text(strip=True) if name_tag else ""
                # parse code
                code = int(seat_str)
                span = sec.find("span", string=lambda t: t and "code" in t.lower())
                if span:
                    m = re.search(r"(\d+)", span.get_text())
                    if m:
                        code = int(m.group(1))
                # parse table
                table = sec.find("table")
                if not table:
                    return None
                data = {"name": name, "code": code}
                for tr in table.find_all("tr"):
                    cols = [td.get_text(strip=True) for td in tr.find_all("td")]
                    if len(cols) == 2:
                        subj = cols[0].split("/")[0].strip()
                        data[subj] = cols[1]
                return data
            except Exception as e:
                logging.error(f"{seat_str} → {e}")
                await asyncio.sleep(0.5)
        logging.error(f"{seat_str} → failed after retries")
        return None

# --- 3) Run tasks with progress bar ---
async def run():
    async with aiohttp.ClientSession() as session:
        seats = [
            str(s).zfill(6)
            for s in range(START_SEAT, END_SEAT + 1)
            if str(s).zfill(6) not in seen_codes
        ]
        tasks = [fetch_one(session, seat) for seat in seats]
        for coro in tqdm(asyncio.as_completed(tasks),
                         total=len(tasks), desc="Fetching"):
            item = await coro
            if item:
                code_str = str(item['code']).zfill(6)
                print(f"[OK]   {code_str} → {item['name']}")
                results.append(item)

if __name__ == "__main__":
    # run scraper
    asyncio.run(run())

    # --- 4) Combine old + new, rewrite XLSX ---
    all_records = existing + results
    if not all_records:
        print("No new records fetched.")
        exit()

    # gather all headers
    hdrs = set().union(*(r.keys() for r in all_records))
    headers = ["name", "code"] + sorted(h for h in hdrs if h not in ("name", "code"))

    wb = xlsxwriter.Workbook(OUTPUT_XLSX)
    ws = wb.add_worksheet()
    # write header
    for c, h in enumerate(headers):
        ws.write(0, c, h)
    # write rows
    for r_idx, rec in enumerate(all_records, start=1):
        for c, h in enumerate(headers):
            ws.write(r_idx, c, rec.get(h, ""))

    wb.close()
    print(f"\nWrote {len(all_records)} total records to {OUTPUT_XLSX}")
    if os.path.exists(ERROR_LOG):
        print(f"Check {ERROR_LOG} for any failures.")
